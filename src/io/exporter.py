"""将内存结果输出到 Excel（4张Sheet）。"""
from __future__ import annotations
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..models import Month, Order, Staff
from ..engine.rules import RulesConfig
from ..engine.scheduler import ScheduleResult
from ..utils import month_range, window_range
from ..reporter import order_summary, schedule_table, transfer_table, warnings as warn_mod
from ..reporter.schedule_table import BORROW_MARKER

log = logging.getLogger(__name__)

_YELLOW_FONT = Font(color="FF8C00")      # 黄色字体（用于借还行姓名）
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")

# 订单情况表：每个订单组循环使用的背景色（白/浅蓝/浅绿/浅橙/浅紫）
_ORDER_GROUP_FILLS = [
    PatternFill("solid", fgColor="FFFFFF"),
    PatternFill("solid", fgColor="DAEEF3"),
    PatternFill("solid", fgColor="E2EFDA"),
    PatternFill("solid", fgColor="FCE4D6"),
    PatternFill("solid", fgColor="EBE7F2"),
]


def export(
    result: ScheduleResult,
    rules: RulesConfig,
    start_month: Month,
    end_month: Month,
    output_path: str,
    loader_warnings: List[str],
    calendar=None,
    data_date=None,
    end_date=None,
) -> str:
    months = month_range(start_month, end_month)
    # 若 scheduler 已生成 windows 则复用，否则按 (data_date, end_date) 现场生成
    windows = result.windows if result.windows else (
        window_range(calendar, start_month, end_month, data_date, end_date)
        if calendar is not None else None
    )
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_order_summary(wb, result.orders, months, rules, data_date)
    _write_schedule_table(wb, result.staff_pool, result.orders, windows)
    _write_transfer_table(wb, result.staff_pool)
    _write_warnings(wb, result.orders, result.staff_pool, rules, months,
                    loader_warnings, result.warnings)

    wb.save(output_path)
    log.info("Output saved: %s", output_path)
    return output_path


def _write_order_summary(wb, orders: Dict[str, Order], months: List[Month],
                         rules: RulesConfig, data_date=None) -> None:
    ws = wb.create_sheet("各订单排班后的情况表")
    header, rows, order_groups = order_summary.build_rows(
        orders, months, rules.work_days_per_month, data_date)
    _write_sheet(ws, header, rows)

    n_cols = len(header)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, (_, start_row, end_row) in enumerate(order_groups):
        fill = _ORDER_GROUP_FILLS[idx % len(_ORDER_GROUP_FILLS)]
        # Apply background color to all cells in this group
        for r in range(start_row, end_row + 1):
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = fill

        # Merge 项目编号 (col B=2) and 合同名称 (col C=3) when multiple level rows
        if end_row > start_row:
            ws.merge_cells(f"B{start_row}:B{end_row}")
            ws.merge_cells(f"C{start_row}:C{end_row}")
            ws.cell(row=start_row, column=2).alignment = center_wrap
            ws.cell(row=start_row, column=3).alignment = center_wrap


def _write_schedule_table(wb, staff_pool: Dict[str, Staff], orders: Dict[str, Order],
                          windows) -> None:
    ws = wb.create_sheet("最新排班表")
    header, rows = schedule_table.build_rows(staff_pool, orders, windows)
    _write_sheet(ws, header, rows)

    # 处理借还行黄色字体 + 空行（BORROW_MARKER 已混入 staff_name 列）
    name_col = 2   # B列（1-based）
    for r_idx, row in enumerate(rows, start=2):
        if row[0] is None:   # 空行
            continue
        raw_name = row[1]
        if isinstance(raw_name, str) and raw_name.startswith(BORROW_MARKER):
            clean = raw_name[len(BORROW_MARKER):]
            cell = ws.cell(row=r_idx, column=name_col)
            cell.value = clean
            cell.font = _YELLOW_FONT


def _write_transfer_table(wb, staff_pool: Dict[str, Staff]) -> None:
    ws = wb.create_sheet("人员转场日期表")
    header, rows = transfer_table.build_rows(staff_pool)
    _write_sheet(ws, header, rows)
    # 转场日期列格式化
    date_col = 5
    for row in ws.iter_rows(min_row=2, min_col=date_col, max_col=date_col):
        for cell in row:
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = 'YYYY"年"M"月"D"日"'


def _write_warnings(wb, orders: Dict[str, Order], staff_pool: Dict[str, Staff],
                    rules: RulesConfig, months: List[Month],
                    loader_warnings: List[str], engine_warnings: List[str]) -> None:
    ws = wb.create_sheet("预警说明")
    header, rows = warn_mod.build_rows(orders, staff_pool, rules, months,
                                       loader_warnings, engine_warnings)
    _write_sheet(ws, header, rows)


def _write_sheet(ws, header: List[str], rows: List[list]) -> None:
    ws.append(header)
    _style_header(ws)
    for row in rows:
        ws.append([None if v is None and row[0] is None else v for v in row])
    _auto_col_width(ws)
    ws.freeze_panes = "A2"


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_col_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
