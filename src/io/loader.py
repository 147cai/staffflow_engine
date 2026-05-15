"""读取输入Excel，构建内部模型对象。"""
from __future__ import annotations
import re
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

import openpyxl
import yaml

from ..models import Assignment, BorrowConfig, Month, Order, Staff, WorkCalendar

log = logging.getLogger(__name__)


@dataclass
class LoadResult:
    orders: Dict[str, Order]
    staff_pool: Dict[str, Staff]
    calendar: WorkCalendar
    borrow_configs: List[BorrowConfig]
    warnings: List[str] = field(default_factory=list)


def load_all(excel_path: str, schema_path: str, calendar_path: str) -> LoadResult:
    schema = _load_yaml(schema_path)
    cal = _load_calendar(calendar_path)
    level_map = schema.get("level_map", {})
    warnings: List[str] = []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    orders = _load_orders(wb, schema["orders"], level_map, warnings)
    _load_workload(wb, schema["workload"], orders, level_map, warnings)
    staff_pool = _load_staff_pool(wb, schema["staff_pool"], warnings)
    _load_initial_assignments(wb, schema["initial_assignments"], staff_pool, orders, level_map, warnings)
    _load_whitelists(wb, schema["whitelist"], orders, staff_pool, level_map, warnings)
    borrows = _load_borrow_config(wb, schema["borrow_config"], orders, staff_pool, cal, warnings)

    wb.close()
    return LoadResult(orders=orders, staff_pool=staff_pool, calendar=cal,
                      borrow_configs=borrows, warnings=warnings)


# ── private helpers ──────────────────────────────────────────────────────────

def _load_yaml(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _load_calendar(path: str) -> WorkCalendar:
    raw = _load_yaml(path)

    # 支持两种格式：
    # 新格式：顶层有 monthly / holidays / makeup_days 三个key
    # 旧格式：所有key均为"YYYY-MM"月份字符串（平铺）
    if "monthly" in raw:
        monthly_raw = raw.get("monthly") or {}
        holiday_list = raw.get("holidays") or []
        makeup_list = raw.get("makeup_days") or []
    else:
        monthly_raw = {k: v for k, v in raw.items() if isinstance(k, str) and len(k) == 7}
        holiday_list = []
        makeup_list = []

    work_days = {date(int(k[:4]), int(k[5:7]), 1): int(v)
                 for k, v in monthly_raw.items() if k}

    def _to_date(val) -> date:
        if isinstance(val, date):
            return val
        return date.fromisoformat(str(val))

    holidays = {_to_date(d) for d in holiday_list}
    makeup_days = {_to_date(d) for d in makeup_list}

    return WorkCalendar(work_days=work_days, holidays=holidays, makeup_days=makeup_days)


def _build_col_map(ws, header_rows: int) -> Dict[str, int]:
    """Merge multiple header rows; row-N value overrides row-(N-1) when non-empty."""
    merged: List[Optional[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=header_rows, values_only=True):
        for i, cell in enumerate(row):
            val = str(cell).strip() if cell is not None else None
            if len(merged) <= i:
                merged.append(val)
            elif val:
                merged[i] = val
    return {name: i for i, name in enumerate(merged) if name}


def _get(row: tuple, col_map: Dict[str, int], key: str):
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _str(val) -> Optional[str]:
    return str(val).strip() if val is not None else None


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        from openpyxl.utils.datetime import from_excel
        return from_excel(int(val)).date()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {val!r}")


def _parse_month(val) -> Optional[Month]:
    d = _parse_date(val)
    return date(d.year, d.month, 1) if d else None


def _normalize_level(raw, level_map: dict) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    return level_map.get(s, s) or None


def _clean_order_ref(raw: str) -> str:
    """Remove '（信贷）' annotations and whitespace from borrow config order ref."""
    return re.sub(r'[（(][^）)]*[）)]', '', str(raw)).strip()


def _resolve_order_no(ref: str, orders: Dict[str, Order]) -> Optional[str]:
    """Match short order code (e.g. F150053) to full order number."""
    clean = _clean_order_ref(ref)
    if clean in orders:
        return clean
    for full_no in orders:
        short = full_no.split("-")[-1]
        if short == clean:
            return full_no
    return None


def _load_orders(wb, cfg: dict, level_map: dict, warnings: List[str]) -> Dict[str, Order]:
    ws = wb[cfg["name"]]
    col = _build_col_map(ws, cfg["header_rows"])
    c = cfg["columns"]
    orders: Dict[str, Order] = {}

    for row in ws.iter_rows(min_row=cfg["data_start_row"], values_only=True):
        no = _str(_get(row, col, c["order_no"]))
        if not no:
            continue
        status = _str(_get(row, col, c["status"])) or ""
        if status != "进行中":
            continue
        try:
            budget = {
                "高级": float(_get(row, col, c["senior_budget"]) or 0),
                "中级": float(_get(row, col, c["mid_budget"]) or 0),
                "初级": float(_get(row, col, c["junior_budget"]) or 0),
            }
            orders[no] = Order(
                no=no,
                name=_str(_get(row, col, c["name"])) or no,
                domain=_str(_get(row, col, c["domain"])) or "",
                start_date=_parse_date(_get(row, col, c["start_date"])),
                end_date=_parse_date(_get(row, col, c["end_date"])),
                status=status,
                initial_budget=budget,
            )
        except Exception as e:
            warnings.append(f"[订单清单] 跳过行（{no}）：{e}")
    return orders


def _load_workload(wb, cfg: dict, orders: Dict[str, Order],
                   level_map: dict, warnings: List[str]) -> None:
    ws = wb[cfg["name"]]
    col = _build_col_map(ws, cfg["header_rows"])
    c = cfg["columns"]
    idx_order = cfg["col_order_no"]

    for row in ws.iter_rows(min_row=cfg["data_start_row"], values_only=True):
        if not row or len(row) <= idx_order:
            continue
        no = _str(row[idx_order])
        if not no:
            continue
        if no not in orders:
            warnings.append(f"[订单当前工作量] 订单 {no} 不在订单清单中，已忽略")
            continue
        level = _normalize_level(_get(row, col, c["level"]), level_map)
        raw_consumed = _get(row, col, c["consumed"])
        if level and raw_consumed is not None:
            try:
                consumed = float(raw_consumed)
                orders[no].consumed_before[level] = round(
                    orders[no].consumed_before.get(level, 0) + consumed, 2)
            except (TypeError, ValueError):
                pass

    for order in orders.values():
        order.remaining = {
            lvl: round(order.initial_budget.get(lvl, 0) - order.consumed_before.get(lvl, 0), 2)
            for lvl in order.initial_budget
        }


def _load_staff_pool(wb, cfg: dict, warnings: List[str]) -> Dict[str, Staff]:
    ws = wb[cfg["name"]]
    col = _build_col_map(ws, cfg["header_rows"])
    c = cfg["columns"]
    pool: Dict[str, Staff] = {}

    for row in ws.iter_rows(min_row=cfg["data_start_row"], values_only=True):
        name = _str(_get(row, col, c["name"]))
        domain = _str(_get(row, col, c["domain"]))
        if not name:
            continue
        if name in pool:
            warnings.append(f"[当前在岗人员清单] 重名人员 {name}，仅保留第一条")
            continue
        if not domain:
            warnings.append(f"[当前在岗人员清单] {name} 领域为空，已跳过")
            continue
        pool[name] = Staff(name=name, domain=domain, base_level="")
    return pool


def _load_initial_assignments(wb, cfg: dict, staff_pool: Dict[str, Staff],
                               orders: Dict[str, Order], level_map: dict,
                               warnings: List[str]) -> None:
    ws = wb[cfg["name"]]
    col = _build_col_map(ws, cfg["header_rows"])
    c = cfg["columns"]
    today = date.today()

    for row in ws.iter_rows(min_row=cfg["data_start_row"], values_only=True):
        order_no = _str(_get(row, col, c["order_no"]))
        name = _str(_get(row, col, c["name"]))
        if not name or not order_no:
            continue
        if name not in staff_pool:
            warnings.append(f"[当前订单人员清单] {name} 不在人员池，已忽略")
            continue
        level = _normalize_level(_get(row, col, c["level"]), level_map)
        entry = _parse_date(_get(row, col, c["entry_date"]))
        exit_ = _parse_date(_get(row, col, c["exit_date"]))

        staff = staff_pool[name]
        if not staff.base_level and level:
            staff.base_level = level

        # entry_date取最早的一条（多订单时保留最初进场日）
        if entry and (staff.entry_date is None or entry < staff.entry_date):
            staff.entry_date = entry
        if exit_:
            if staff.exit_date is None or exit_ > staff.exit_date:
                staff.exit_date = exit_

        # 以退场时间是否已过判断当前订单
        if exit_ is None or exit_ >= today:
            staff.initial_order_no = order_no
            staff.current_order_no = order_no
            staff.initial_level = level
            staff.current_level = level


def _load_whitelists(wb, cfg: dict, orders: Dict[str, Order],
                     staff_pool: Dict[str, Staff], level_map: dict,
                     warnings: List[str]) -> None:
    prefix = cfg["prefix"]
    col_name_key = cfg["columns"]["name"]
    col_level_key = cfg["columns"]["level"]
    header_rows = cfg["header_rows"]
    data_start = cfg["data_start_row"]

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith(prefix):
            continue
        raw_suffix = sheet_name[len(prefix):].lstrip("_").strip()
        full_no = _resolve_order_no(raw_suffix, orders)
        if not full_no:
            warnings.append(f"[遴选清单] Sheet '{sheet_name}' 无法匹配订单编号 '{raw_suffix}'，已忽略")
            continue
        ws = wb[sheet_name]
        col = _build_col_map(ws, header_rows)
        whitelist: Dict[str, str] = {}
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            name = _str(_get(row, col, col_name_key))
            level = _normalize_level(_get(row, col, col_level_key), level_map)
            if not name:
                continue
            if name not in staff_pool:
                warnings.append(f"[遴选清单 {sheet_name}] {name} 不在人员池，已告警")
            if level:
                whitelist[name] = level
        orders[full_no].whitelist = whitelist or {}


def _load_borrow_config(wb, cfg: dict, orders: Dict[str, Order],
                        staff_pool: Dict[str, Staff], calendar: WorkCalendar,
                        warnings: List[str]) -> List[BorrowConfig]:
    if cfg["name"] not in wb.sheetnames:
        return []
    ws = wb[cfg["name"]]
    col = _build_col_map(ws, cfg["header_rows"])
    c = cfg["columns"]
    results: List[BorrowConfig] = []

    for row in ws.iter_rows(min_row=cfg["data_start_row"], values_only=True):
        name = _str(_get(row, col, c["name"]))
        from_domain = _str(_get(row, col, c["from_domain"]))
        to_order_raw = _str(_get(row, col, c["to_order"]))
        month_raw = _get(row, col, c["month"])
        days_raw = _get(row, col, c["days"])

        if not name or not to_order_raw or days_raw is None:
            continue
        try:
            days = int(float(str(days_raw)))
        except (TypeError, ValueError):
            warnings.append(f"[借还配置] {name} 借用天数无效 '{days_raw}'，已跳过")
            continue
        if days <= 0:
            warnings.append(f"[借还配置] {name} 借用天数 {days} ≤ 0，已跳过")
            continue

        month = _parse_month(month_raw)
        if not month:
            warnings.append(f"[借还配置] {name} 发生月份无法解析 '{month_raw}'，已跳过")
            continue

        to_order_no = _resolve_order_no(to_order_raw, orders)
        if not to_order_no:
            warnings.append(f"[借还配置] {name} 借入订单 '{to_order_raw}' 不在订单清单，已跳过")
            continue

        if name not in staff_pool:
            warnings.append(f"[借还配置] {name} 不在人员池，已跳过")
            continue

        # Split across consecutive months if days exceeds one month's work days
        current_month = month
        remaining = days
        while remaining > 0:
            month_days = calendar.get_working_days(current_month)
            if month_days <= 0:
                warnings.append(
                    f"[借还配置] {name} {current_month} 日历中无工作日数据，借还截止（剩余 {remaining} 天未安排）")
                break
            alloc = min(remaining, month_days)
            results.append(BorrowConfig(
                staff_name=name,
                from_domain=from_domain or staff_pool[name].domain,
                to_order_no=to_order_no,
                month=current_month,
                days=alloc,
            ))
            remaining -= alloc
            if current_month.month == 12:
                current_month = date(current_month.year + 1, 1, 1)
            else:
                current_month = date(current_month.year, current_month.month + 1, 1)
    return results
